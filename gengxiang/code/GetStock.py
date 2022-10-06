import json
import time
from urllib import request

import openpyxl
import pymysql

# 今天日期
today = time.strftime('%Y-%m-%d', time.localtime(time.time()))


# http获取当前行情信息
def get_current(stock_code):
    current = request.urlopen('http://qt.gtimg.cn/q=' + stock_code).read().decode('gbk')
    # print(current)
    cur_arr = str(current).split('~')
    stock = {
        't_date': cur_arr[30][0:4] + '-' + cur_arr[30][4:6] + '-' + cur_arr[30][6:8],
        'name': cur_arr[1],
        'code': stock_code,
        'price': float(cur_arr[3]),
        'volume': int(float(cur_arr[6])),
        't_volume': int(float(cur_arr[37])),
        't_change': float(cur_arr[38])
    }
    return stock


# http获取历史数据
def get_history(stock_code, start_date, end_date):
    history = request.urlopen('https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?param=' + stock_code + ',day,'
                              + start_date + ',' + end_date + ',260,qfq').read().decode('gbk')
    if history.find('"code":0') == -1:
        history = request.urlopen('https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?param=' + stock_code + ',day,'
                                  + start_date + ',' + end_date + ',260,qfq').read().decode('gbk')
    # print(history)
    history_stock = json.loads(history)['data'][stock_code]
    stock_name = history_stock['qt'][stock_code][1]
    result_list = []
    for h_stock in history_stock['day']:
        stock = {
            't_date': h_stock[0],
            'name': stock_name,
            'code': stock_code,
            'price': float(h_stock[2]),
            'volume': int(float(h_stock[5])),
            't_volume': 0,
            't_change': 0
        }
        result_list.append(stock)
    return result_list


# excel 保存
def save_excel(stock_list, file_name):
    excel = openpyxl.load_workbook(file_name + '.xlsx')
    sheet = excel[excel.sheetnames[0]]
    i = 2
    for history in stock_list:
        sheet.cell(row=i, column=1).value = history['t_date']
        sheet.cell(row=i, column=2).value = history['name']
        sheet.cell(row=i, column=3).value = history['code']
        sheet.cell(row=i, column=4).value = history['price']
        sheet.cell(row=i, column=5).value = history['volume']
        sheet.cell(row=i, column=6).value = history['t_volume']
        sheet.cell(row=i, column=7).value = history['t_change']
        i = i + 1
    excel.save(file_name + today + '.xlsx')


# excel 读取
def get_dict_from_refer(file_name):
    # 打开Excel表格
    refer_excel = openpyxl.load_workbook(file_name)
    # 获取指定Sheet表单页
    refer_sheet = refer_excel[refer_excel.sheetnames[0]]
    # 创建字典：创建一个以型号为key，以品牌为value的字典
    stock_list = []
    # 行循环：从第二行开始循环，到最后一行截止
    for row in range(2, refer_sheet.max_row + 1):
        # 读取cell单元格中的数据
        stock = {
            't_date': str((refer_sheet.cell(row=row, column=1)).value),
            'name': (refer_sheet.cell(row=row, column=2)).value,
            'code': (refer_sheet.cell(row=row, column=3)).value,
            'price': float((refer_sheet.cell(row=row, column=4)).value),
            'volume': int(float((refer_sheet.cell(row=row, column=5)).value)),
            't_volume': int(float((refer_sheet.cell(row=row, column=6)).value)),
            't_change': float((refer_sheet.cell(row=row, column=7)).value)
        }
        stock_list.append(stock)
    return stock_list


def save_mysql(stock_list):
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    for stock_info in stock_list:
        sql = "SELECT * FROM gx_stock_data WHERE code = '%s' and t_date = '%s' limit 1" % \
              (stock_info['code'], stock_info['t_date'])
        cursor.execute(sql)
        results = cursor.fetchall()
        if len(results) == 0:
            sql = "INSERT INTO gx_stock_data(code, name, price, volume, t_volume, t_change, t_date)" \
                  "VALUES ('%s', '%s', %s, %s, %s, %s, '%s')" % \
                  (stock_info['code'], stock_info['name'], stock_info['price'], stock_info['volume'],
                   stock_info['t_volume'], stock_info['t_change'], stock_info['t_date'])
            cursor.execute(sql)
        else:
            sql = "UPDATE gx_stock_data SET price = %s, volume= %s, t_volume= %s, t_change = %s  " \
                  "WHERE code = '%s' and t_date = '%s'" % \
                  (stock_info['price'], stock_info['volume'],
                   stock_info['t_volume'], stock_info['t_change'], stock_info['code'], stock_info['t_date'])
            cursor.execute(sql)
    mysql.commit()
    cursor.close()


def get_mysql(stock_code):
    stock_list = []
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    sql = "SELECT * FROM gx_stock_data WHERE code = '%s' order by id desc limit 20" % stock_code
    cursor.execute(sql)
    results = cursor.fetchall()
    for row in results:
        stock = {
            't_date': row[7],
            'name': row[2],
            'code': row[1],
            'price': row[3],
            'volume': row[4],
            't_volume': row[5],
            't_change': row[6]
        }
        stock_list.append(stock)
    return stock_list


# 获取基础信息
today = get_current('sh000001')
print(today)

# 获取历史信息
# history_list = get_history('sh000001', '2022-09-01', '2022-10-10')
history_list = [today]
print(history_list)

"""
# 保存excel
save_excel(history_list, 'E://StockTips//gengxiang//data//sh000001')
# 读取excel
history_list = get_dict_from_refer('E://StockTips//gengxiang//data//sh000001' + today + '.xlsx')
"""

save_mysql(history_list)
print(get_mysql('sh000001'))
