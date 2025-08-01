import json
import os
import time
from urllib import request

import openpyxl
import pymysql

# 今天日期
todayStr = time.strftime('%Y-%m-%d', time.localtime(time.time()))


# http获取当前行情信息
def get_current(stock_code):
    current = request.urlopen('http://qt.gtimg.cn/q=' + stock_code).read().decode('gbk')
    print(current)
    cur_arr = str(current).split('~')
    stock = {
        't_date': cur_arr[30][0:4] + '-' + cur_arr[30][4:6] + '-' + cur_arr[30][6:8],
        'name': cur_arr[1],
        'code': stock_code,
        'price': float(cur_arr[3]),
        'volume': int(float(cur_arr[6])),
        't_volume': int(float(cur_arr[37])),
        't_change': float(cur_arr[38])
    }
    return stock


# http获取历史数据
def get_history(stock_code, start_date, end_date):
    history = request.urlopen('https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?param=' + stock_code + ',day,'
                              + start_date + ',' + end_date + ',260,qfq').read().decode('gbk')
    if history.find('"code":0') == -1:
        history = request.urlopen('https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?param=' + stock_code + ',day,'
                                  + start_date + ',' + end_date + ',260,qfq').read().decode('gbk')
    # print(history)
    history_stock = json.loads(history)['data'][stock_code]
    stock_name = history_stock['qt'][stock_code][1]
    result_list = []
    for h_stock in history_stock['day']:
        stock = {
            't_date': h_stock[0],
            'name': stock_name,
            'code': stock_code,
            'price': float(h_stock[2]),
            'volume': int(float(h_stock[5])),
            't_volume': 0,
            't_change': 0
        }
        result_list.append(stock)
    return result_list


# excel 保存
def save_excel(stock_list, file_name):
    if not os.path.isfile(file_name):
        openpyxl.Workbook().save(file_name)

    excel = openpyxl.load_workbook(file_name)
    sheet = excel[excel.sheetnames[0]]
    sheet.cell(row=1, column=1).value = '日期'
    sheet.cell(row=1, column=2).value = '名称'
    sheet.cell(row=1, column=3).value = '代码'
    sheet.cell(row=1, column=4).value = '价格'
    sheet.cell(row=1, column=5).value = '涨停价格'
    sheet.cell(row=1, column=6).value = '成交额(万元)'
    sheet.cell(row=1, column=7).value = '涨跌幅%'
    sheet.cell(row=1, column=8).value = '量比'
    sheet.cell(row=1, column=9).value = '总市值(亿元)'
    sheet.cell(row=1, column=10).value = '换手率'
    sheet.cell(row=1, column=11).value = '市盈率'
    sheet.cell(row=1, column=12).value = '市净率'
    i = 2
    for history in stock_list:
        sheet.cell(row=i, column=1).value = history['date']
        sheet.cell(row=i, column=2).value = history['name']
        sheet.cell(row=i, column=3).value = history['code']
        sheet.cell(row=i, column=4).value = history['price']
        sheet.cell(row=i, column=5).value = history['stop_price']
        sheet.cell(row=i, column=6).value = history['amo']
        sheet.cell(row=i, column=7).value = history['amp']
        sheet.cell(row=i, column=8).value = history['qrr']
        sheet.cell(row=i, column=9).value = history['mc']
        sheet.cell(row=i, column=10).value = history['hs']
        sheet.cell(row=i, column=11).value = history['per']
        sheet.cell(row=i, column=12).value = history['pb']
        i = i + 1

    excel.save(file_name)


# excel 读取
def get_excel(file_name):
    # 打开Excel表格
    refer_excel = openpyxl.load_workbook(file_name)
    # 获取指定Sheet表单页
    refer_sheet = refer_excel[refer_excel.sheetnames[0]]
    # 创建字典：创建一个以型号为key，以品牌为value的字典
    stock_list = []
    # 行循环：从第二行开始循环，到最后一行截止
    for row in range(2, refer_sheet.max_row + 1):
        # 读取cell单元格中的数据
        stock = {
            'date': str((refer_sheet.cell(row=row, column=1)).value),
            'name': (refer_sheet.cell(row=row, column=2)).value,
            'code': (refer_sheet.cell(row=row, column=3)).value,
            'price': float((refer_sheet.cell(row=row, column=4)).value),
            'stop_price': float((refer_sheet.cell(row=row, column=5)).value),
            'amo': int(float((refer_sheet.cell(row=row, column=6)).value)),
            'amp': float((refer_sheet.cell(row=row, column=7)).value),
            'qrr': float((refer_sheet.cell(row=row, column=8)).value),
            'mc': float((refer_sheet.cell(row=row, column=9)).value),
            'hs': float((refer_sheet.cell(row=row, column=10)).value),
            'per': float((refer_sheet.cell(row=row, column=11)).value),
            'pb': float((refer_sheet.cell(row=row, column=12)).value)
        }
        stock_list.append(stock)
    return stock_list


def save_mysql(stock_list):
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    for stock_info in stock_list:
        sql = "SELECT * FROM stock_data_detail WHERE code = '%s' and date = '%s' limit 1" % \
              (stock_info['code'], stock_info['date'])
        cursor.execute(sql)
        results = cursor.fetchall()
        if len(results) == 0:
            sql = "INSERT INTO stock_data_detail(code, name, date, price, stop_price, b_stop_price, amo, amp, qrr, mc, hs, per, pb)" \
                  "VALUES ('%s', '%s', '%s', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)" % \
                  (stock_info['code'], stock_info['name'], stock_info['date'], stock_info['price'],
                   stock_info['stop_price'], stock_info['b_stop_price'],
                   stock_info['amo'], stock_info['amp'], stock_info['qrr'], stock_info['mc'], stock_info['hs'],
                   stock_info['per'], stock_info['pb'])
            cursor.execute(sql)
        else:
            sql = "UPDATE stock_data_detail SET price = %s, amo= %s,b_stop_price= %s, amp= %s, qrr = %s, mc = %s, hs = %s, per = %s, pb = %s " \
                  "WHERE code = '%s' and date = '%s'" % \
                  (stock_info['price'], stock_info['amo'], stock_info['b_stop_price'], stock_info['amp'],
                   stock_info['qrr'], stock_info['mc'],
                   stock_info['hs'], stock_info['per'], stock_info['pb'], stock_info['code'], stock_info['date'])
            cursor.execute(sql)
    mysql.commit()
    cursor.close()


def get_mysql(stock_code):
    stock_list = []
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    sql = "SELECT * FROM stock_data_detail WHERE code = '%s' order by id desc limit 30" % stock_code
    cursor.execute(sql)
    results = cursor.fetchall()
    for row in results:
        stock = {
            'date': row[3],
            'name': row[2],
            'code': row[1],
            'price': row[4],
            'b_stop_price': row[5],  # 跌涨停价
            'stop_price': row[6],  # 涨停价
            'amo': row[8],  # 成交额（万元）
            'amp': row[7],  # 涨跌幅%
            'qrr': row[10],  # 量比
            'mc': row[9],  # 总市值
            'hs': row[11],  # 换手率
            'per': row[12],  # 市盈率
            'pb': row[13],  # 市净率
        }
        stock_list.append(stock)
    return stock_list


def get_stop_mysql():
    stock_list = []
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    sql = "SELECT a.top_industry, count( b.date ) AS stop_times, GROUP_CONCAT( DISTINCT a.stock_name ) AS stop_details FROM stock_info a RIGHT JOIN stock_data_detail b ON a.all_code = b.`code` WHERE date > DATE_SUB( CURRENT_DATE, INTERVAL 30 DAY ) AND price = stop_price GROUP BY a.top_industry ORDER BY stop_times DESC"
    cursor.execute(sql)
    results = cursor.fetchall()
    for row in results:
        stock = {
            'industry': row[0],
            'stop_times': row[1],
            'stop_details': row[2].split(',')
        }
        stock_list.append(stock)
    return stock_list


def get_second1_stop_mysql():
    stock_list = []
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    sql = "SELECT a.second_industry, count( b.date ) AS stop_times, GROUP_CONCAT( DISTINCT a.stock_name ) AS stop_details " \
          "FROM stock_info a RIGHT JOIN stock_data_detail b ON a.all_code = b.`code` WHERE date > DATE_SUB( CURRENT_DATE, INTERVAL 2 DAY ) AND price = stop_price GROUP BY a.second_industry ORDER BY stop_times DESC"
    cursor.execute(sql)
    results = cursor.fetchall()
    for row in results:
        stock = {
            'industry': row[0],
            'stop_times': row[1],
            'stop_details': row[2].split(',')
        }
        stock_list.append(stock)
    return stock_list


def get_second2_stop_mysql():
    stock_list = []
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    sql = "SELECT a.second_industry, count( b.date ) AS stop_times, GROUP_CONCAT( DISTINCT a.stock_name ) AS stop_details FROM stock_info a RIGHT JOIN stock_data_detail b ON a.all_code = b.`code` WHERE date > DATE_SUB( CURRENT_DATE, INTERVAL 15 DAY ) and date <= (select DISTINCT date from stock_data_detail GROUP BY date order by date desc LIMIT 1, 1) AND price = stop_price GROUP BY a.second_industry ORDER BY stop_times DESC"
    cursor.execute(sql)
    results = cursor.fetchall()
    for row in results:
        stock = {
            'industry': row[0],
            'stop_times': row[1],
            'stop_details': row[2].split(',')
        }
        stock_list.append(stock)
    return stock_list


def get_today_stop_mysql():
    stock_list = []
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    sql = "SELECT a.second_industry, count( b.date ) AS stop_times, GROUP_CONCAT( DISTINCT a.stock_name ) AS stop_details FROM stock_info a RIGHT JOIN stock_data_detail b ON a.all_code = b.`code` WHERE date = (select DISTINCT date from stock_data_detail GROUP BY date order by date desc LIMIT 0, 1) AND price = stop_price GROUP BY a.second_industry ORDER BY stop_times DESC"
    cursor.execute(sql)
    results = cursor.fetchall()
    for row in results:
        stock = {
            'industry': row[0],
            'stop_times': row[1],
            'stop_details': row[2].split(',')
        }
        stock_list.append(stock)
    return stock_list


def get_stop_rank_mysql():
    stock_list = []
    mysql = pymysql.connect(host='127.0.0.1', port=3366, user='root', password='gengxiang',
                            database='stock_tips', charset='utf8')
    cursor = mysql.cursor()
    sql = "SELECT a.CODE, a.stock_name, a.industry, a.address, a.concept, count( b.date ) AS stop_times FROM stock_info a RIGHT JOIN stock_data_detail b ON a.all_code = b.`code`  WHERE date > (select DISTINCT date from stock_data_detail GROUP BY date order by date desc LIMIT 8, 1)  AND price = stop_price GROUP BY b.CODE ORDER BY stop_times DESC"
    cursor.execute(sql)
    results = cursor.fetchall()
    for row in results:
        stock = {
            'code': row[0],
            'name': row[1],
            'industry': row[2],
            'address': row[3],
            'times': row[5],
            'concept': row[4].split(';')
        }
        stock_list.append(stock)
    return stock_list


# http获取当前行情信息
def get_current_batch(stock_codes, write_file):
    # current_stree = request.urlopen('https://gateway.jrj.com/quot-feed/category_hqs', timeout=1.0).read().decode('gbk')
    # print(current_stree)

    stocks = []
    try:
        current_str = request.urlopen('http://qt.gtimg.cn/q=' + ','.join(stock_codes), timeout=5.0).read().decode('gbk')
    except:
        print("请求异常等待………………")
        time.sleep(5)
        current_str = request.urlopen('http://qt.gtimg.cn/q=' + ','.join(stock_codes), timeout=5.0).read().decode('gbk')

    if write_file:
        with open("..\data\\" + todayStr + ".txt", "a") as file:
            file.write(current_str)
            print("写入文件->", current_str)
    else:
        currents = current_str.split(';')
        print(len(currents), "->", currents)
        for ccs in range(0, len(currents)):
            if '\n' != currents[ccs]:
                current_arr = str(currents[ccs]).split('~')
                if current_arr[39] == '':
                    current_arr[39] = 0
                if current_arr[38] == '':
                    continue

                stock = {
                    'date': current_arr[30][0:4] + '-' + current_arr[30][4:6] + '-' + current_arr[30][6:8],  # 时间
                    'name': current_arr[1],  # 名称
                    'code': stock_codes[ccs],  # 编码
                    'price': float(current_arr[3]),  # 收盘价格
                    'stop_price': float(current_arr[47]),  # 涨停价格
                    'b_stop_price': float(current_arr[48]),  # 跌停价格
                    'amo': int(float(current_arr[37])),  # 成交额（万元）
                    'amp': float(current_arr[32]),  # 涨跌幅%
                    'qrr': float(current_arr[49]),  # 量比
                    'hs': float(current_arr[38]),  # 换手率
                    'mc': float(current_arr[45]),  # 总市值
                    'per': float(current_arr[39]),  # 市盈率
                    'pb': float(current_arr[46]),  # 市净率
                }
                stocks.append(stock)
    return stocks

# current_str = request.urlopen('https://gateway.jrj.com/quot-feed/category_hqs', timeout=1.0).read().decode('gbk')
